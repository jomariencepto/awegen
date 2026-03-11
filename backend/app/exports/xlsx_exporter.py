import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from app.utils.logger import get_logger

logger = get_logger(__name__)

# ── Colour palette ────────────────────────────────────────────────────────────
BLUE_DARK   = "1E3A5F"   # header background
BLUE_MID    = "2563EB"   # section header background
BLUE_LIGHT  = "DBEAFE"   # alternate row tint
WHITE       = "FFFFFF"
GRAY_LIGHT  = "F3F4F6"
GREEN_BG    = "D1FAE5"
GREEN_DARK  = "065F46"
RED_TEXT    = "DC2626"
PURPLE_BG   = "EDE9FE"
PURPLE_DARK = "5B21B6"
YELLOW_BG   = "FEF9C3"
YELLOW_DARK = "92400E"

# ── Shared style helpers ──────────────────────────────────────────────────────

def _thin_border():
    s = Side(style='thin', color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, value="", bold=False, size=10,
          bg=None, fg="000000", wrap=True, align="left", italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=fg, italic=italic,
                  name="Calibri")
    c.alignment = Alignment(wrap_text=wrap, vertical="top",
                            horizontal=align)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.border = _thin_border()
    return c

def _merge(ws, r1, c1, r2, c2, value="", bold=False, size=11,
           bg=None, fg="000000", align="center"):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = Font(bold=bold, size=size, color=fg, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=True)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


# ── School header block ───────────────────────────────────────────────────────

def _write_school_header(ws, exam_data, total_cols):
    """Rows 1-3: school header spanning all columns."""
    _merge(ws, 1, 1, 1, total_cols,
           value="PAMBAYANG DALUBHASAAN NG MARILAO",
           bold=True, size=13, bg=BLUE_DARK, fg=WHITE)

    _merge(ws, 2, 1, 2, total_cols,
           value="College of Computer Studies  |  Information Technology Department",
           bold=False, size=10, bg=BLUE_DARK, fg="BFDBFE")

    exam_type = exam_data.get('category_name', exam_data.get('category', 'EXAMINATION'))
    _merge(ws, 3, 1, 3, total_cols,
           value=f"{str(exam_type).upper()} — {str(exam_data.get('title', '')).upper()}",
           bold=True, size=11, bg=BLUE_MID, fg=WHITE)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 18
    return 4   # next available row


def _write_exam_info(ws, exam_data, start_row, total_cols):
    """A small info strip below the header."""
    teacher_name = str(exam_data.get('teacher_name') or 'N/A')
    subject_name = str(
        exam_data.get('subject_name')
        or exam_data.get('subject')
        or exam_data.get('module_title')
        or 'N/A'
    )
    info = (
        f"Created by: {teacher_name}  |  "
        f"Subject: {subject_name}  |  "
        f"Duration: {exam_data.get('duration_minutes', 60)} min  |  "
        f"Total Questions: {exam_data.get('total_questions', '?')}  |  "
        f"Passing Score: {exam_data.get('passing_score', '?')}"
    )
    _merge(ws, start_row, 1, start_row, total_cols,
           value=info, bold=False, size=9, bg=GRAY_LIGHT, fg="374151",
           align="left")
    ws.row_dimensions[start_row].height = 16
    return start_row + 1


# ── Question-type section writers ─────────────────────────────────────────────

def _section_header(ws, row, total_cols, roman, label, bg=BLUE_MID):
    _merge(ws, row, 1, row, total_cols,
           value=f"{roman}.  {label}",
           bold=True, size=10, bg=bg, fg=WHITE, align="left")
    ws.row_dimensions[row].height = 16
    return row + 1


def _col_header_row(ws, row, headers, widths, bgs=None):
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        bg = (bgs[i-1] if bgs else BLUE_DARK)
        c = _cell(ws, row, i, h, bold=True, size=9,
                  bg=bg, fg=WHITE, align="center", wrap=False)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 14
    return row + 1


# ── Multiple Choice ───────────────────────────────────────────────────────────

def _write_mc(ws, questions, start_row, is_answer_key=False):
    total_cols = 7
    start_row = _section_header(
        ws, start_row, total_cols, "I",
        "MULTIPLE CHOICE: Choose the best answer.")

    headers = ["#", "Question", "A", "B", "C", "D",
               "Answer" if is_answer_key else ""]
    widths  = [4, 45, 18, 18, 18, 18, 12]
    start_row = _col_header_row(ws, start_row, headers, widths)

    for i, q in enumerate(questions):
        row = start_row + i
        opts = q.get('options', [])
        correct = str(q.get('correct_answer', '')).strip()

        tint = BLUE_LIGHT if i % 2 == 0 else WHITE

        _cell(ws, row, 1, i + 1, align="center", bg=tint, size=9)
        _cell(ws, row, 2, q.get('question_text', ''), bg=tint, size=9)
        for j, opt in enumerate(opts[:4], 3):
            letter = chr(64 + j - 2)   # A=3→A, B=4→B …
            is_correct = (opt == correct or correct == letter)
            cell_bg = GREEN_BG if (is_answer_key and is_correct) else tint
            cell_fg = GREEN_DARK if (is_answer_key and is_correct) else "000000"
            _cell(ws, row, j, opt, bg=cell_bg, fg=cell_fg, size=9,
                  bold=(is_answer_key and is_correct))
        if is_answer_key:
            _cell(ws, row, 7, correct, bold=True,
                  bg=GREEN_BG, fg=GREEN_DARK, align="center", size=9)
        ws.row_dimensions[row].height = 30

    return start_row + len(questions)


# ── True / False ──────────────────────────────────────────────────────────────

def _write_tf(ws, questions, start_row, is_answer_key=False, roman="II"):
    total_cols = 7
    start_row = _section_header(
        ws, start_row, total_cols, roman,
        "TRUE OR FALSE: Write True if the statement is correct, otherwise False.")

    headers = ["#", "Statement",
               "Answer" if is_answer_key else "True / False"]
    widths  = [4, 75, 14]
    # pad remaining cols
    for col in range(4, 8):
        ws.column_dimensions[get_column_letter(col)].width = 1

    start_row = _col_header_row(ws, start_row, headers, widths)

    for i, q in enumerate(questions):
        row = start_row + i
        correct = str(q.get('correct_answer', 'True')).strip()
        tint = BLUE_LIGHT if i % 2 == 0 else WHITE
        _cell(ws, row, 1, i + 1, align="center", bg=tint, size=9)
        _cell(ws, row, 2, q.get('question_text', ''), bg=tint, size=9)
        ans_bg = GREEN_BG if is_answer_key else tint
        ans_fg = GREEN_DARK if is_answer_key else "000000"
        _cell(ws, row, 3,
              correct if is_answer_key else "",
              bold=is_answer_key, bg=ans_bg, fg=ans_fg,
              align="center", size=9)
        ws.row_dimensions[row].height = 28

    return start_row + len(questions)


# ── Fill in the Blank ─────────────────────────────────────────────────────────

def _write_fib(ws, questions, start_row, is_answer_key=False, roman="III"):
    total_cols = 7
    start_row = _section_header(
        ws, start_row, total_cols, roman,
        "FILL IN THE BLANK: Write the correct word or phrase on the space provided.")

    headers = ["#", "Statement / Sentence",
               "Answer" if is_answer_key else "Write Answer Here"]
    widths  = [4, 70, 20]
    start_row = _col_header_row(ws, start_row, headers, widths)

    for i, q in enumerate(questions):
        row = start_row + i
        correct = str(q.get('correct_answer', '')).strip()
        text    = str(q.get('question_text', '')).strip()
        tint = BLUE_LIGHT if i % 2 == 0 else WHITE

        _cell(ws, row, 1, i + 1, align="center", bg=tint, size=9)
        _cell(ws, row, 2, text, bg=tint, size=9)
        _cell(ws, row, 3,
              correct if is_answer_key else "",
              bold=is_answer_key,
              bg=GREEN_BG if is_answer_key else tint,
              fg=GREEN_DARK if is_answer_key else "000000",
              align="center", size=9)
        ws.row_dimensions[row].height = 28

    return start_row + len(questions)


# ── Identification ────────────────────────────────────────────────────────────

def _write_identification(ws, questions, start_row,
                          is_answer_key=False, roman="IV"):
    total_cols = 7
    start_row = _section_header(
        ws, start_row, total_cols, roman,
        "IDENTIFICATION: Identify the term or concept being described.")

    headers = ["#", "Description",
               "Answer" if is_answer_key else "Write Answer Here"]
    widths  = [4, 70, 20]
    start_row = _col_header_row(ws, start_row, headers, widths)

    for i, q in enumerate(questions):
        row = start_row + i
        correct = str(q.get('correct_answer', '')).strip()
        tint = BLUE_LIGHT if i % 2 == 0 else WHITE

        _cell(ws, row, 1, i + 1, align="center", bg=tint, size=9)
        _cell(ws, row, 2, q.get('question_text', ''), bg=tint, size=9)
        _cell(ws, row, 3,
              correct if is_answer_key else "",
              bold=is_answer_key,
              bg=GREEN_BG if is_answer_key else tint,
              fg=GREEN_DARK if is_answer_key else "000000",
              align="center", size=9)
        ws.row_dimensions[row].height = 28

    return start_row + len(questions)


# ── Shared sheet builder ──────────────────────────────────────────────────────

SECTION_ORDER = ['multiple_choice', 'true_false', 'fill_in_blank', 'identification']
ROMAN         = ['I', 'II', 'III', 'IV']
SECTION_LABELS = {
    'multiple_choice': "MULTIPLE CHOICE: Choose the best answer.",
    'true_false':      "TRUE OR FALSE: Write True if the statement is correct, otherwise False.",
    'fill_in_blank':   "FILL IN THE BLANK: Write the correct word or phrase on the space provided.",
    'identification':  "IDENTIFICATION: Identify the term or concept being described.",
}

def _build_exam_sheet(ws, exam_data, is_answer_key=False,
                      include_header=True, is_special=False):
    """Write a full exam / answer-key into worksheet *ws*."""
    total_cols = 7
    ws.sheet_view.showGridLines = False

    questions = exam_data.get('questions', [])
    grouped   = {}
    for q in questions:
        qt = q.get('question_type', 'multiple_choice')
        grouped.setdefault(qt, []).append(q)

    # ── Title tab colour ──────────────────────────────────────────────────────
    if is_answer_key:
        ws.sheet_properties.tabColor = "16A34A"
    elif is_special:
        ws.sheet_properties.tabColor = "CA8A04"
    else:
        ws.sheet_properties.tabColor = "2563EB"

    # ── School / exam header ──────────────────────────────────────────────────
    if include_header:
        row = _write_school_header(ws, exam_data, total_cols)
        row = _write_exam_info(ws, exam_data, row, total_cols)
    else:
        row = 1

    # ── Answer-key label ──────────────────────────────────────────────────────
    if is_answer_key:
        _merge(ws, row, 1, row, total_cols,
               value="★  ANSWER KEY — FOR TEACHER USE ONLY  ★",
               bold=True, size=10, bg="FEF2F2", fg=RED_TEXT)
        ws.row_dimensions[row].height = 14
        row += 1

    if is_special:
        _merge(ws, row, 1, row, total_cols,
               value="⚡  SPECIAL EXAM — QUESTIONS RANDOMIZED",
               bold=True, size=10, bg=YELLOW_BG, fg=YELLOW_DARK)
        ws.row_dimensions[row].height = 14
        row += 1

    # ── Spacer ────────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 6
    row += 1

    # ── Question sections ─────────────────────────────────────────────────────
    roman_idx = 0
    for qt in SECTION_ORDER:
        if qt not in grouped:
            continue
        qs = grouped[qt]
        roman = ROMAN[roman_idx]
        roman_idx += 1

        if qt == 'multiple_choice':
            row = _write_mc(ws, qs, row, is_answer_key)
        elif qt == 'true_false':
            row = _write_tf(ws, qs, row, is_answer_key, roman)
        elif qt == 'fill_in_blank':
            row = _write_fib(ws, qs, row, is_answer_key, roman)
        elif qt == 'identification':
            row = _write_identification(ws, qs, row, is_answer_key, roman)

        # spacer between sections
        ws.row_dimensions[row].height = 8
        row += 1

    # ── Freeze panes below header ─────────────────────────────────────────────
    freeze_row = 6 if include_header else 1
    ws.freeze_panes = f"A{freeze_row}"

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1


# ══════════════════════════════════════════════════════════════════════════════
# Public XLSXExporter class
# ══════════════════════════════════════════════════════════════════════════════

class XLSXExporter:
    """Export exams and answer keys to formatted .xlsx files."""

    # ── Regular exam ──────────────────────────────────────────────────────────
    def export_exam(self, exam_data, output_path,
                    include_header=True, is_special=False):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Special Exam" if is_special else "Exam"
            _build_exam_sheet(ws, exam_data,
                              is_answer_key=False,
                              include_header=include_header,
                              is_special=is_special)
            wb.save(output_path)
            logger.info(f"✅ XLSX exam saved: {output_path}")
            return True
        except Exception as e:
            logger.error(f"❌ XLSX exam error: {e}", exc_info=True)
            return False

    # ── TOS ───────────────────────────────────────────────────────────────────
    def export_tos(self, tos_data, output_path):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Table of Specifications"
            ws.sheet_view.showGridLines = False
            ws.sheet_properties.tabColor = "7C3AED"

            total_cols = 6

            # ── Title header ─────────────────────────────────────────────────
            _merge(ws, 1, 1, 1, total_cols,
                   value="PAMBAYANG DALUBHASAAN NG MARILAO",
                   bold=True, size=13, bg=BLUE_DARK, fg=WHITE)
            _merge(ws, 2, 1, 2, total_cols,
                   value="College of Computer Studies  |  Information Technology Department",
                   bold=False, size=10, bg=BLUE_DARK, fg="BFDBFE")
            _merge(ws, 3, 1, 3, total_cols,
                   value="TABLE OF SPECIFICATIONS",
                   bold=True, size=12, bg=BLUE_MID, fg=WHITE)
            _merge(ws, 4, 1, 4, total_cols,
                   value=tos_data.get('exam_title', 'Untitled Exam').upper(),
                   bold=True, size=11, bg="1E40AF", fg=WHITE)

            info = (
                f"Total Questions: {tos_data.get('total_questions', 0)}  |  "
                f"Duration: {tos_data.get('duration_minutes', 60)} min"
            )
            _merge(ws, 5, 1, 5, total_cols, value=info,
                   bold=False, size=9, bg=GRAY_LIGHT, fg="374151", align="left")

            for r in range(1, 6):
                ws.row_dimensions[r].height = 18 if r <= 4 else 13

            row = 6
            ws.row_dimensions[row].height = 8
            row += 1

            # ── Cognitive Distribution ────────────────────────────────────────
            _merge(ws, row, 1, row, total_cols,
                   value="COGNITIVE LEVEL DISTRIBUTION",
                   bold=True, size=10, bg=BLUE_MID, fg=WHITE, align="left")
            ws.row_dimensions[row].height = 15
            row += 1

            for col, (h, w) in enumerate(zip(
                ["Cognitive Level", "Questions", "Percentage"],
                [28, 14, 14]
            ), 1):
                _cell(ws, row, col, h, bold=True, size=9,
                      bg=BLUE_DARK, fg=WHITE, align="center", wrap=False)
                ws.column_dimensions[get_column_letter(col)].width = w
            ws.row_dimensions[row].height = 13
            row += 1

            cognitive_dist = tos_data.get('cognitive_distribution', {})
            cognitive_pct  = tos_data.get('cognitive_percentages', {})
            for i, (level, count) in enumerate(cognitive_dist.items()):
                tint = BLUE_LIGHT if i % 2 == 0 else WHITE
                pct  = cognitive_pct.get(level, 0)
                _cell(ws, row, 1, level.replace('_', ' ').title(), bg=tint, size=9)
                _cell(ws, row, 2, count, align="center", bg=tint, size=9)
                _cell(ws, row, 3, f"{pct}%", align="center", bg=tint, size=9)
                ws.row_dimensions[row].height = 13
                row += 1

            ws.row_dimensions[row].height = 8
            row += 1

            # ── Difficulty Distribution ───────────────────────────────────────
            _merge(ws, row, 1, row, total_cols,
                   value="DIFFICULTY LEVEL DISTRIBUTION",
                   bold=True, size=10, bg="065F46", fg=WHITE, align="left")
            ws.row_dimensions[row].height = 15
            row += 1

            for col, (h, w) in enumerate(zip(
                ["Difficulty Level", "Questions", "Percentage"],
                [28, 14, 14]
            ), 1):
                _cell(ws, row, col, h, bold=True, size=9,
                      bg="064E3B", fg=WHITE, align="center", wrap=False)
                ws.column_dimensions[get_column_letter(col)].width = w
            ws.row_dimensions[row].height = 13
            row += 1

            difficulty_dist = tos_data.get('difficulty_distribution', {})
            difficulty_pct  = tos_data.get('difficulty_percentages', {})
            for i, (level, count) in enumerate(difficulty_dist.items()):
                tint = GREEN_BG if i % 2 == 0 else WHITE
                pct  = difficulty_pct.get(level, 0)
                _cell(ws, row, 1, level.title(), bg=tint, size=9)
                _cell(ws, row, 2, count, align="center", bg=tint, size=9)
                _cell(ws, row, 3, f"{pct}%", align="center", bg=tint, size=9)
                ws.row_dimensions[row].height = 13
                row += 1

            ws.row_dimensions[row].height = 8
            row += 1

            # ── Topic-Cognitive Matrix ────────────────────────────────────────
            matrix = tos_data.get('topic_cognitive_matrix', {})
            if matrix and cognitive_dist:
                _merge(ws, row, 1, row, total_cols,
                       value="TOPIC — COGNITIVE LEVEL MATRIX",
                       bold=True, size=10, bg=PURPLE_DARK, fg=WHITE, align="left")
                ws.row_dimensions[row].height = 15
                row += 1

                cog_levels = list(cognitive_dist.keys())
                matrix_headers = ["Topic"] + [l.replace('_', ' ').title() for l in cog_levels] + ["Total"]
                matrix_widths  = [32] + [14] * len(cog_levels) + [10]

                for col, (h, w) in enumerate(zip(matrix_headers, matrix_widths), 1):
                    _cell(ws, row, col, h, bold=True, size=9,
                          bg=PURPLE_DARK, fg=WHITE, align="center", wrap=False)
                    ws.column_dimensions[get_column_letter(col)].width = w
                ws.row_dimensions[row].height = 13
                row += 1

                for i, (topic, levels) in enumerate(matrix.items()):
                    tint = PURPLE_BG if i % 2 == 0 else WHITE
                    _cell(ws, row, 1, topic, bg=tint, size=9)
                    row_total = 0
                    for col_i, level in enumerate(cog_levels, 2):
                        val = levels.get(level, 0)
                        row_total += val
                        _cell(ws, row, col_i, val, align="center", bg=tint, size=9)
                    _cell(ws, row, len(cog_levels) + 2, row_total,
                          align="center", bold=True, bg=tint, size=9)
                    ws.row_dimensions[row].height = 13
                    row += 1

            ws.freeze_panes = "A7"
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToPage  = True
            ws.page_setup.fitToWidth = 1

            wb.save(output_path)
            logger.info(f"✅ XLSX TOS saved: {output_path}")
            return True
        except Exception as e:
            logger.error(f"❌ XLSX TOS error: {e}", exc_info=True)
            return False

    # ── Answer key ────────────────────────────────────────────────────────────
    def export_answer_key(self, answer_key_data, output_path,
                          include_header=True, is_special=False):
        try:
            # Normalise data shape (mirrors word_exporter logic)
            if 'questions' in answer_key_data:
                exam_data = answer_key_data
            elif 'answer_key' in answer_key_data:
                exam_data = answer_key_data['answer_key']
            else:
                exam_data = answer_key_data

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Special Answer Key" if is_special else "Answer Key"
            _build_exam_sheet(ws, exam_data,
                              is_answer_key=True,
                              include_header=include_header,
                              is_special=is_special)
            wb.save(output_path)
            logger.info(f"✅ XLSX answer key saved: {output_path}")
            return True
        except Exception as e:
            logger.error(f"❌ XLSX answer key error: {e}", exc_info=True)
            return False
